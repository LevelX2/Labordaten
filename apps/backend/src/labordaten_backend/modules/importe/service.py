from __future__ import annotations

import csv
import hashlib
import io
import json
import re
import unicodedata
from dataclasses import dataclass
from datetime import date, datetime
from pathlib import Path

from openpyxl import load_workbook
from sqlalchemy import func, select
from sqlalchemy.orm import Session

from labordaten_backend.core.labor_value_formatting import (
    DEFAULT_LOWER_REFERENCE_OPERATOR,
    DEFAULT_UPPER_REFERENCE_OPERATOR,
)
from labordaten_backend.core.documents import get_documents_root, store_document_file, store_existing_document_path
from labordaten_backend.models.base import utcnow
from labordaten_backend.models.befund import Befund
from labordaten_backend.models.dokument import Dokument
from labordaten_backend.models.einheit import Einheit
from labordaten_backend.models.einheit_alias import EinheitAlias
from labordaten_backend.models.gruppen_parameter import GruppenParameter
from labordaten_backend.models.import_pruefpunkt import ImportPruefpunkt
from labordaten_backend.models.importvorgang import Importvorgang
from labordaten_backend.models.labor import Labor
from labordaten_backend.models.laborparameter import Laborparameter
from labordaten_backend.models.laborparameter_alias import LaborparameterAlias
from labordaten_backend.models.messwert import Messwert
from labordaten_backend.models.messwert_referenz import MesswertReferenz
from labordaten_backend.models.parameter_gruppe import ParameterGruppe
from labordaten_backend.models.person import Person
from labordaten_backend.modules.einheiten import service as einheiten_service
from labordaten_backend.modules.gruppen import schemas as gruppen_schemas
from labordaten_backend.modules.gruppen import service as gruppen_service
from labordaten_backend.modules.importe.schemas import (
    ImportAehnlicheGruppeRead,
    ImportBefundPreviewRead,
    ImportEntwurfCreate,
    ImportGruppenvorschlaegeAnwendenRequest,
    ImportGruppenvorschlaegeAnwendenResponse,
    ImportGruppenvorschlagErgebnisRead,
    ImportKomplettEntfernenRead,
    ImportMesswertPayload,
    ImportMesswertPreviewRead,
    ImportParameterMapping,
    ImportParameterVorschlagPayload,
    ImportParameterVorschlagRead,
    ImportPayload,
    ImportPromptCreate,
    ImportPromptRead,
    ImportPruefpunktRead,
    ImportUebernehmenRequest,
    ImportvorgangDetailRead,
    ImportvorgangListRead,
    ImportGruppenvorschlagRead,
)
from labordaten_backend.modules.parameter.normalization import build_parameter_key_candidate, normalize_parameter_name
from labordaten_backend.modules.parameter import conversions as parameter_conversions
from labordaten_backend.modules.parameter import service as parameter_service


HEADER_ALIASES: dict[str, set[str]] = {
    "person_id": {"personid", "person_id"},
    "labor_id": {"laborid", "labor_id"},
    "labor_name": {"laborname", "labor_name", "labor"},
    "entnahmedatum": {"entnahmedatum", "entnahme", "sampledate", "datum"},
    "befunddatum": {"befunddatum", "reportdate"},
    "befund_bemerkung": {"befundbemerkung", "befund_bemerkung", "berichtsbemerkung"},
    "gruppe_name": {"gruppe", "gruppenname", "gruppe_name", "block", "blockname", "kategorie"},
    "gruppe_beschreibung": {"gruppenbeschreibung", "gruppe_beschreibung", "blockbeschreibung"},
    "parameter_id": {"parameterid", "parameter_id"},
    "original_parametername": {
        "originalparametername",
        "parametername",
        "parameter",
        "name",
        "wertname",
        "analyte",
    },
    "wert_typ": {"werttyp", "wert_typ", "typ"},
    "wert_operator": {"wertoperator", "wert_operator", "operator"},
    "wert_roh_text": {"wertrohtext", "wert_roh_text", "wert", "value", "result"},
    "wert_num": {"wertnum", "wert_num", "zahlwert", "numerisch"},
    "wert_text": {"werttext", "wert_text", "textwert"},
    "einheit_original": {"einheitoriginal", "einheit_original", "einheit", "unit"},
    "bemerkung_kurz": {"bemerkungkurz", "bemerkung_kurz", "bemerkung"},
    "bemerkung_lang": {"bemerkunglang", "bemerkung_lang"},
    "ki_hinweis": {"kihinweis", "ki_hinweis", "extraktionshinweis", "extraktions_hinweis"},
    "referenz_text_original": {"referenztextoriginal", "referenztext", "referenz"},
    "untere_grenze_num": {"unteregrenzenum", "untere_grenze_num", "referenzuntere", "untergrenze"},
    "untere_grenze_operator": {"unteregrenzeoperator", "untere_grenze_operator", "referenzuntereoperator"},
    "obere_grenze_num": {"oberegrenzenum", "obere_grenze_num", "referenzobere", "obergrenze"},
    "obere_grenze_operator": {"oberegrenzeoperator", "obere_grenze_operator", "referenzobereoperator"},
    "referenz_einheit": {"referenzeinheit", "referenz_einheit"},
    "referenz_geschlecht_code": {"referenzgeschlechtcode", "referenz_geschlecht_code", "geschlechtreferenz"},
    "referenz_alter_min_tage": {"referenzaltermintage", "referenz_alter_min_tage"},
    "referenz_alter_max_tage": {"referenzaltermaxtage", "referenz_alter_max_tage"},
    "referenz_bemerkung": {"referenzbemerkung", "referenz_bemerkung"},
    "alias_uebernehmen": {"aliasuebernehmen", "alias_uebernehmen", "aliasanlegen"},
    "unsicher_flag": {"unsicherflag", "unsicher_flag", "unsicher"},
    "pruefbedarf_flag": {"pruefbedarfflag", "pruefbedarf_flag", "pruefbedarf"},
}


@dataclass
class Pruefregel:
    objekt_typ: str
    objekt_schluessel_temp: str | None
    pruefart: str
    status: str
    meldung: str


@dataclass
class ParameterResolution:
    parameter_id: str | None
    herkunft: str | None
    hinweis: str | None = None
    mehrdeutig: bool = False


@dataclass
class AliasResolution:
    requested: bool
    alias_text: str | None
    conflict_message: str | None = None
    already_present: bool = False


@dataclass
class ExistingGroupContext:
    gruppe_id: str
    name: str
    parameter_ids: set[str]
    parameter_names_by_id: dict[str, str]


class ParameterResolver:
    def __init__(self, db: Session) -> None:
        self.by_id: dict[str, Laborparameter] = {}
        self._schluessel_lookup: dict[str, list[tuple[str, str]]] = {}
        self._anzeigename_lookup: dict[str, list[tuple[str, str]]] = {}
        self._alias_lookup: dict[str, list[tuple[str, str]]] = {}

        for parameter in db.scalars(select(Laborparameter)):
            self.by_id[parameter.id] = parameter
            self._add_candidate(
                self._schluessel_lookup,
                normalize_parameter_name(parameter.interner_schluessel),
                parameter.id,
                parameter.interner_schluessel,
            )
            self._add_candidate(
                self._anzeigename_lookup,
                normalize_parameter_name(parameter.anzeigename),
                parameter.id,
                parameter.anzeigename,
            )

        for alias in db.scalars(select(LaborparameterAlias)):
            self._add_candidate(
                self._alias_lookup,
                alias.alias_normalisiert,
                alias.laborparameter_id,
                alias.alias_text,
            )

    def resolve(
        self,
        *,
        original_name: str,
        explicit_parameter_id: str | None,
        manual_parameter_id: str | None,
    ) -> ParameterResolution:
        if manual_parameter_id:
            return ParameterResolution(parameter_id=manual_parameter_id, herkunft="manuell")
        if explicit_parameter_id:
            return ParameterResolution(parameter_id=explicit_parameter_id, herkunft="explizit")

        normalized_name = normalize_parameter_name(original_name)
        if not normalized_name:
            return ParameterResolution(parameter_id=None, herkunft=None)

        for herkunft, lookup in (
            ("schluessel", self._schluessel_lookup),
            ("anzeigename", self._anzeigename_lookup),
            ("alias", self._alias_lookup),
        ):
            resolution = self._match_lookup(lookup, normalized_name, herkunft)
            if resolution is not None:
                return resolution

        return ParameterResolution(parameter_id=None, herkunft=None)

    def get_parameter_name(self, parameter_id: str | None) -> str | None:
        if not parameter_id:
            return None
        parameter = self.by_id.get(parameter_id)
        return parameter.anzeigename if parameter is not None else None

    def _match_lookup(
        self,
        lookup: dict[str, list[tuple[str, str]]],
        normalized_name: str,
        herkunft: str,
    ) -> ParameterResolution | None:
        candidates = lookup.get(normalized_name, [])
        unique_ids = {parameter_id for parameter_id, _ in candidates}
        if not unique_ids:
            return None
        if len(unique_ids) > 1:
            namen = ", ".join(
                sorted({self.by_id[parameter_id].anzeigename for parameter_id in unique_ids if parameter_id in self.by_id})
            )
            return ParameterResolution(
                parameter_id=None,
                herkunft=herkunft,
                hinweis=namen or None,
                mehrdeutig=True,
            )

        parameter_id = next(iter(unique_ids))
        matched_texts = sorted({label for candidate_id, label in candidates if candidate_id == parameter_id})
        return ParameterResolution(
            parameter_id=parameter_id,
            herkunft=herkunft,
            hinweis=matched_texts[0] if matched_texts else self.get_parameter_name(parameter_id),
        )

    @staticmethod
    def _add_candidate(
        lookup: dict[str, list[tuple[str, str]]],
        normalized_name: str,
        parameter_id: str,
        label: str,
    ) -> None:
        if not normalized_name:
            return
        lookup.setdefault(normalized_name, []).append((parameter_id, label))


def list_importe(db: Session) -> list[ImportvorgangListRead]:
    stmt = select(Importvorgang).order_by(Importvorgang.erstellt_am.desc())
    return [_build_list_item(db, importvorgang) for importvorgang in db.scalars(stmt)]


def get_import_detail(db: Session, import_id: str) -> ImportvorgangDetailRead | None:
    importvorgang = db.get(Importvorgang, import_id)
    if importvorgang is None:
        return None
    return _build_detail(db, importvorgang)


def create_import_prompt(db: Session, payload: ImportPromptCreate) -> ImportPromptRead:
    context = {
        "bekannteLabore": _build_prompt_labor_context(db),
        "bekannteParameter": _build_prompt_parameter_context(db),
        "bekannteEinheiten": _build_prompt_unit_context(db),
        "bekannteGruppen": _build_prompt_group_context(db),
    }
    context_json = json.dumps(context, ensure_ascii=False, indent=2, sort_keys=True)
    summary = (
        f"Labore: {len(context['bekannteLabore'])}; "
        f"Parameter: {len(context['bekannteParameter'])}; "
        f"Einheiten: {len(context['bekannteEinheiten'])}; "
        f"Gruppen: {len(context['bekannteGruppen'])}; "
        f"Prompt: {_format_prompt_type(payload.prompt_typ)}"
    )
    return ImportPromptRead(
        promptText=_build_import_prompt_text(context_json, prompt_typ=payload.prompt_typ),
        kontextZusammenfassung=summary,
        schemaVersion="1.0",
    )


def list_pruefpunkte(db: Session, import_id: str) -> list[ImportPruefpunkt]:
    stmt = (
        select(ImportPruefpunkt)
        .where(ImportPruefpunkt.importvorgang_id == import_id)
        .order_by(ImportPruefpunkt.status.desc(), ImportPruefpunkt.objekt_typ.asc())
    )
    return list(db.scalars(stmt))


def create_import_entwurf(db: Session, payload: ImportEntwurfCreate) -> ImportvorgangDetailRead:
    import_payload = _parse_payload(payload.payload_json, payload.person_id_override)
    dokument_id = payload.dokument_id
    if dokument_id is None and import_payload.befund.dokument_pfad:
        dokument = store_existing_document_path(
            source_path=import_payload.befund.dokument_pfad,
            dokument_typ="importquelle",
            originalquelle_behalten=True,
            bemerkung=payload.bemerkung,
        )
        db.add(dokument)
        db.flush()
        dokument_id = dokument.id

    return _create_import_entwurf_record(
        db,
        import_payload=import_payload,
        bemerkung=payload.bemerkung,
        dokument_id=dokument_id,
    )


def create_import_entwurf_from_json_upload(
    db: Session,
    *,
    payload_json: str,
    person_id_override: str | None,
    import_bemerkung: str | None,
    document_filename: str | None,
    document_content_type: str | None,
    document_content: bytes | None,
    document_name_override: str | None,
) -> ImportvorgangDetailRead:
    import_payload = _parse_payload(payload_json)
    import_payload.befund.person_id = person_id_override
    dokument_id: str | None = None

    if document_content:
        dokument = store_document_file(
            content=document_content,
            original_filename=_build_import_document_filename(
                db,
                payload=import_payload,
                original_filename=document_filename or "Laborbericht",
                override_name=document_name_override,
            ),
            content_type=document_content_type,
            dokument_typ="importquelle",
            originalquelle_behalten=True,
            bemerkung=import_bemerkung,
        )
        db.add(dokument)
        db.flush()
        dokument_id = dokument.id
    elif import_payload.befund.dokument_pfad:
        dokument = store_existing_document_path(
            source_path=import_payload.befund.dokument_pfad,
            dokument_typ="importquelle",
            originalquelle_behalten=True,
            bemerkung=import_bemerkung,
        )
        db.add(dokument)
        db.flush()
        dokument_id = dokument.id

    return _create_import_entwurf_record(
        db,
        import_payload=import_payload,
        bemerkung=import_bemerkung,
        dokument_id=dokument_id,
    )


def create_import_entwurf_from_file(
    db: Session,
    *,
    filename: str,
    content_type: str | None,
    content: bytes,
    person_id_override: str | None,
    labor_id_override: str | None,
    labor_name_override: str | None,
    entnahmedatum_override: date | None,
    befunddatum_override: date | None,
    befund_bemerkung_override: str | None,
    import_bemerkung: str | None,
    quelle_behalten: bool,
) -> ImportvorgangDetailRead:
    dokument: Dokument | None = None
    if quelle_behalten:
        dokument = store_document_file(
            content=content,
            original_filename=filename,
            content_type=content_type,
            dokument_typ="importquelle",
            originalquelle_behalten=True,
            bemerkung=import_bemerkung,
        )
        db.add(dokument)
        db.flush()

    import_payload = _parse_tabular_import(
        filename=filename,
        content=content,
        person_id_override=person_id_override,
        labor_id_override=labor_id_override,
        labor_name_override=labor_name_override,
        entnahmedatum_override=entnahmedatum_override,
        befunddatum_override=befunddatum_override,
        befund_bemerkung_override=befund_bemerkung_override,
        dokument=dokument,
    )
    return _create_import_entwurf_record(
        db,
        import_payload=import_payload,
        bemerkung=import_bemerkung,
        dokument_id=dokument.id if dokument is not None else None,
    )


def uebernehmen_import(
    db: Session,
    import_id: str,
    payload: ImportUebernehmenRequest,
) -> ImportvorgangDetailRead:
    importvorgang = db.get(Importvorgang, import_id)
    if importvorgang is None:
        raise ValueError("Importvorgang nicht gefunden.")
    if importvorgang.status == "uebernommen":
        raise ValueError("Dieser Importvorgang wurde bereits übernommen.")

    import_payload = _parse_payload(importvorgang.roh_payload_text or "", payload.person_id_override)
    if payload.person_id_override:
        importvorgang.roh_payload_text = _serialize_payload(import_payload)
        importvorgang.person_id_vorschlag = import_payload.befund.person_id
        importvorgang.fingerprint = hashlib.sha256(importvorgang.roh_payload_text.encode("utf-8")).hexdigest()
        db.add(importvorgang)

    mappings = payload.parameter_mappings
    pruefregeln = _generate_checks(db, importvorgang.id, import_payload, mappings)

    fehler = [item for item in pruefregeln if item.status == "fehler"]
    warnungen = [item for item in pruefregeln if item.status == "warnung"]
    if fehler:
        detail = " ".join(item.meldung for item in fehler[:3])
        raise ValueError(f"Der Import enthält noch Fehler und kann noch nicht übernommen werden. {detail}")
    if warnungen and not payload.bestaetige_warnungen:
        raise ValueError("Bitte Warnungen bewusst bestätigen, bevor der Import übernommen wird.")

    person_id = import_payload.befund.person_id
    if person_id is None:
        raise ValueError("Dem Import ist keine Person zugeordnet.")

    labor_id = _resolve_labor(db, import_payload)
    befund = Befund(
        person_id=person_id,
        labor_id=labor_id,
        dokument_id=importvorgang.dokument_id,
        entnahmedatum=import_payload.befund.entnahmedatum,
        befunddatum=import_payload.befund.befunddatum,
        bemerkung=import_payload.befund.bemerkung,
        importvorgang_id=importvorgang.id,
        quelle_typ="import",
        duplikat_warnung=bool(warnungen),
    )
    db.add(befund)
    db.flush()

    mapping_lookup = {item.messwert_index: item for item in mappings}
    parameter_suggestion_by_index = _build_parameter_suggestion_by_measurement_index(import_payload)
    parameter_resolver = ParameterResolver(db)
    created_measurements: list[Messwert] = []

    for index, item in enumerate(import_payload.messwerte):
        mapping_request = mapping_lookup.get(index)
        measurement_unit = (
            einheiten_service.ensure_einheit_exists(db, item.einheit_original)
            if item.wert_typ == "numerisch"
            else None
        )
        reference_unit = (
            einheiten_service.ensure_einheit_exists(db, item.referenz_einheit)
            if item.wert_typ == "numerisch"
            else None
        )
        if _is_new_parameter_mapping(mapping_request):
            parameter_id = _create_parameter_from_import_mapping(
                db,
                item=item,
                mapping_request=mapping_request,
                standard_einheit=measurement_unit,
                parameter_vorschlag=parameter_suggestion_by_index.get(index),
            )
        else:
            resolution = parameter_resolver.resolve(
                original_name=item.original_parametername,
                explicit_parameter_id=item.parameter_id,
                manual_parameter_id=mapping_request.laborparameter_id if mapping_request is not None else None,
            )
            parameter_id = resolution.parameter_id
            if parameter_id is None:
                raise ValueError("Es fehlt noch mindestens eine Parameterzuordnung.")

            alias_resolution = _resolve_alias_request(
                db,
                parameter_id=parameter_id,
                original_name=item.original_parametername,
                alias_requested=item.alias_uebernehmen or bool(mapping_request and mapping_request.alias_uebernehmen),
            )
            if alias_resolution.conflict_message:
                raise ValueError(alias_resolution.conflict_message)
            if alias_resolution.requested and alias_resolution.alias_text and not alias_resolution.already_present:
                _create_import_alias(
                    db,
                    parameter_id=parameter_id,
                    alias_text=alias_resolution.alias_text,
                    bemerkung="Bei Import-Zuordnung übernommen",
                )

        normalized = parameter_conversions.resolve_measurement_normalization(
            db,
            laborparameter_id=parameter_id,
            wert_typ=item.wert_typ,
            wert_num=item.wert_num,
            einheit_original=measurement_unit,
        )

        messwert = Messwert(
            person_id=person_id,
            befund_id=befund.id,
            laborparameter_id=parameter_id,
            original_parametername=item.original_parametername,
            wert_typ=item.wert_typ,
            wert_operator=item.wert_operator,
            wert_roh_text=item.wert_roh_text,
            wert_num=item.wert_num,
            wert_text=item.wert_text if item.wert_typ == "text" else None,
            einheit_original=measurement_unit,
            wert_normiert_num=normalized.wert_normiert_num,
            einheit_normiert=normalized.einheit_normiert,
            umrechnungsregel_id=normalized.umrechnungsregel_id,
            bemerkung_kurz=item.bemerkung_kurz,
            bemerkung_lang=item.bemerkung_lang,
            unsicher_flag=item.unsicher_flag,
            pruefbedarf_flag=item.pruefbedarf_flag,
            importvorgang_id=importvorgang.id,
        )
        db.add(messwert)
        db.flush()
        created_measurements.append(messwert)

        if (
            item.referenz_text_original
            or item.untere_grenze_num is not None
            or item.obere_grenze_num is not None
            or item.referenz_einheit
        ):
            db.add(
                MesswertReferenz(
                    messwert_id=messwert.id,
                    referenz_typ="labor",
                    referenz_text_original=item.referenz_text_original,
                    wert_typ=item.wert_typ,
                    untere_grenze_num=item.untere_grenze_num if item.wert_typ == "numerisch" else None,
                    untere_grenze_operator=item.untere_grenze_operator if item.wert_typ == "numerisch" else None,
                    obere_grenze_num=item.obere_grenze_num if item.wert_typ == "numerisch" else None,
                    obere_grenze_operator=item.obere_grenze_operator if item.wert_typ == "numerisch" else None,
                    einheit=reference_unit,
                    soll_text=item.referenz_text_original if item.wert_typ == "text" else None,
                    geschlecht_code=item.referenz_geschlecht_code,
                    alter_min_tage=item.referenz_alter_min_tage,
                    alter_max_tage=item.referenz_alter_max_tage,
                    bemerkung=item.referenz_bemerkung,
                )
            )

    _replace_checks(db, importvorgang.id, pruefregeln, bestaetige_warnungen=payload.bestaetige_warnungen)
    importvorgang.status = "uebernommen"
    importvorgang.warnungen_text = _summarize_warnings(pruefregeln)

    db.add(importvorgang)
    db.commit()
    db.refresh(importvorgang)
    return _build_detail(db, importvorgang)


def verwerfen_import(db: Session, import_id: str) -> ImportvorgangDetailRead:
    importvorgang = db.get(Importvorgang, import_id)
    if importvorgang is None:
        raise ValueError("Importvorgang nicht gefunden.")
    importvorgang.status = "verworfen"
    db.add(importvorgang)
    db.commit()
    db.refresh(importvorgang)
    return _build_detail(db, importvorgang)


def komplett_entfernen_import(
    db: Session,
    import_id: str,
    *,
    dokument_entfernen: bool,
) -> ImportKomplettEntfernenRead:
    importvorgang = db.get(Importvorgang, import_id)
    if importvorgang is None:
        raise ValueError("Importvorgang nicht gefunden.")
    if importvorgang.status == "uebernommen":
        raise ValueError("Übernommene Importe können nicht komplett entfernt werden. Bitte die Löschprüfung verwenden.")

    messwerte_anzahl = db.scalar(select(func.count(Messwert.id)).where(Messwert.importvorgang_id == import_id)) or 0
    befunde_anzahl = db.scalar(select(func.count(Befund.id)).where(Befund.importvorgang_id == import_id)) or 0
    if messwerte_anzahl or befunde_anzahl:
        raise ValueError("Dieser Import hat bereits übernommene Daten. Bitte die Löschprüfung verwenden.")

    pruefpunkte = list(db.scalars(select(ImportPruefpunkt).where(ImportPruefpunkt.importvorgang_id == import_id)))
    pruefpunkte_entfernt = len(pruefpunkte)
    for pruefpunkt in pruefpunkte:
        db.delete(pruefpunkt)

    dokument_id = importvorgang.dokument_id
    dokument_entfernt = False
    dokument = db.get(Dokument, dokument_id) if dokument_id and dokument_entfernen else None
    if dokument is not None:
        other_imports = (
            db.scalar(
                select(func.count(Importvorgang.id))
                .where(Importvorgang.dokument_id == dokument.id)
                .where(Importvorgang.id != import_id)
            )
            or 0
        )
        befund_refs = db.scalar(select(func.count(Befund.id)).where(Befund.dokument_id == dokument.id)) or 0
        if other_imports or befund_refs:
            raise ValueError("Das Dokument wird noch an anderer Stelle verwendet und kann nicht entfernt werden.")

    db.delete(importvorgang)

    if dokument is not None:
        _delete_stored_document_file(dokument)
        db.delete(dokument)
        dokument_entfernt = True

    db.commit()
    return ImportKomplettEntfernenRead(
        import_id=import_id,
        dokument_id=dokument_id,
        dokument_entfernt=dokument_entfernt,
        pruefpunkte_entfernt=pruefpunkte_entfernt,
    )


def anwenden_gruppenvorschlaege(
    db: Session,
    import_id: str,
    payload: ImportGruppenvorschlaegeAnwendenRequest,
) -> ImportGruppenvorschlaegeAnwendenResponse:
    importvorgang = db.get(Importvorgang, import_id)
    if importvorgang is None:
        raise ValueError("Importvorgang nicht gefunden.")
    if importvorgang.status != "uebernommen":
        raise ValueError("Gruppenvorschläge können erst nach der Übernahme des Imports angewendet werden.")

    import_payload = _parse_payload(importvorgang.roh_payload_text or "")
    if not import_payload.gruppen_vorschlaege:
        raise ValueError("Für diesen Import sind keine Gruppenvorschläge vorhanden.")

    imported_measurements = list(
        db.scalars(
            select(Messwert)
            .where(Messwert.importvorgang_id == import_id)
            .order_by(Messwert.erstellt_am.asc())
        )
    )

    results: list[ImportGruppenvorschlagErgebnisRead] = []
    for item in payload.vorschlaege:
        if item.vorschlag_index < 0 or item.vorschlag_index >= len(import_payload.gruppen_vorschlaege):
            raise ValueError(f"Der Gruppenvorschlag mit Index {item.vorschlag_index} existiert nicht.")

        suggestion = import_payload.gruppen_vorschlaege[item.vorschlag_index]
        parameter_ids, _ = _resolve_group_suggestion_parameters(
            payload=import_payload,
            suggestion_index=item.vorschlag_index,
            imported_measurements=imported_measurements,
            parameter_resolver=ParameterResolver(db),
        )

        if item.aktion == "ignorieren":
            results.append(
                ImportGruppenvorschlagErgebnisRead(
                    vorschlag_index=item.vorschlag_index,
                    aktion=item.aktion,
                )
            )
            continue

        if not parameter_ids:
            raise ValueError(
                f"Der Gruppenvorschlag '{suggestion.name}' kann nicht angewendet werden, weil noch keine Parameter zugeordnet sind."
            )

        if item.aktion == "vorhanden":
            if not item.gruppe_id:
                raise ValueError(f"Für den Gruppenvorschlag '{suggestion.name}' fehlt die Zielgruppe.")
            gruppe = gruppen_service.get_gruppe(db, item.gruppe_id)
            if gruppe is None or not gruppe.aktiv:
                raise ValueError(f"Die Zielgruppe für '{suggestion.name}' existiert nicht.")
        elif item.aktion == "neu":
            zielname = (item.gruppenname or suggestion.name).strip()
            if not zielname:
                raise ValueError("Neue Gruppen brauchen einen Namen.")
            gruppe = _find_active_group_by_name(db, zielname)
            if gruppe is None:
                gruppe = gruppen_service.create_gruppe(
                    db,
                    gruppen_schemas.GruppeCreate(
                        name=zielname,
                        beschreibung=suggestion.beschreibung,
                    ),
                )
        else:
            raise ValueError(f"Die Aktion '{item.aktion}' wird nicht unterstützt.")

        gruppen_service.merge_gruppen_parameter(
            db,
            gruppe.id,
            gruppen_schemas.GruppenParameterAssignRequest(
                eintraege=[
                    gruppen_schemas.GruppenParameterAssignItem(
                        laborparameter_id=parameter_id,
                        sortierung=sortierung,
                    )
                    for sortierung, parameter_id in enumerate(parameter_ids, start=1)
                ]
            ),
        )
        results.append(
            ImportGruppenvorschlagErgebnisRead(
                vorschlag_index=item.vorschlag_index,
                aktion=item.aktion,
                gruppe_id=gruppe.id,
                gruppenname=gruppe.name,
                zugeordnete_parameter_anzahl=len(parameter_ids),
            )
        )

    return ImportGruppenvorschlaegeAnwendenResponse(ergebnisse=results)


def _create_import_entwurf_record(
    db: Session,
    *,
    import_payload: ImportPayload,
    bemerkung: str | None,
    dokument_id: str | None,
) -> ImportvorgangDetailRead:
    payload_json = _serialize_payload(import_payload)
    fingerprint = hashlib.sha256(payload_json.encode("utf-8")).hexdigest()

    importvorgang = Importvorgang(
        quelle_typ=import_payload.quelle_typ,
        status="in_pruefung",
        person_id_vorschlag=import_payload.befund.person_id,
        dokument_id=dokument_id,
        roh_payload_text=payload_json,
        schema_version=import_payload.schema_version,
        fingerprint=fingerprint,
        bemerkung=bemerkung,
    )
    db.add(importvorgang)
    db.flush()

    pruefregeln = _generate_checks(db, importvorgang.id, import_payload, [])
    _store_checks(db, importvorgang.id, pruefregeln)
    importvorgang.warnungen_text = _summarize_warnings(pruefregeln)

    db.add(importvorgang)
    db.commit()
    db.refresh(importvorgang)
    return _build_detail(db, importvorgang)


def _build_list_item(db: Session, importvorgang: Importvorgang) -> ImportvorgangListRead:
    counts = _count_checks(db, importvorgang.id)
    payload = _parse_payload(importvorgang.roh_payload_text or "")
    dokument = db.get(Dokument, importvorgang.dokument_id) if importvorgang.dokument_id else None
    return ImportvorgangListRead(
        id=importvorgang.id,
        quelle_typ=importvorgang.quelle_typ,
        status=importvorgang.status,
        person_id_vorschlag=importvorgang.person_id_vorschlag,
        schema_version=importvorgang.schema_version,
        bemerkung=importvorgang.bemerkung,
        dokument_id=importvorgang.dokument_id,
        dokument_dateiname=dokument.dateiname if dokument is not None else None,
        messwerte_anzahl=len(payload.messwerte),
        fehler_anzahl=counts["fehler"],
        warnung_anzahl=counts["warnung"],
        erstellt_am=importvorgang.erstellt_am,
        geaendert_am=importvorgang.geaendert_am,
    )


def _build_prompt_labor_context(db: Session) -> list[dict[str, str]]:
    stmt = select(Labor).where(Labor.aktiv.is_(True)).order_by(Labor.name.asc())
    return [{"id": labor.id, "name": labor.name} for labor in db.scalars(stmt)]


def _build_prompt_parameter_context(db: Session) -> list[dict[str, object]]:
    aliases_by_parameter: dict[str, list[str]] = {}
    alias_stmt = select(LaborparameterAlias).order_by(LaborparameterAlias.alias_text.asc())
    for alias in db.scalars(alias_stmt):
        aliases_by_parameter.setdefault(alias.laborparameter_id, []).append(alias.alias_text)

    stmt = select(Laborparameter).where(Laborparameter.aktiv.is_(True)).order_by(Laborparameter.anzeigename.asc())
    return [
        {
            "id": parameter.id,
            "anzeigename": parameter.anzeigename,
            "internerSchluessel": parameter.interner_schluessel,
            "standardEinheit": parameter.standard_einheit,
            "wertTypStandard": parameter.wert_typ_standard,
            "primaereKlassifikation": parameter.primaere_klassifikation,
            "aliase": aliases_by_parameter.get(parameter.id, []),
        }
        for parameter in db.scalars(stmt)
    ]


def _build_prompt_unit_context(db: Session) -> list[dict[str, object]]:
    aliases_by_unit: dict[str, list[str]] = {}
    alias_stmt = select(EinheitAlias).order_by(EinheitAlias.alias_text.asc())
    for alias in db.scalars(alias_stmt):
        aliases_by_unit.setdefault(alias.einheit_id, []).append(alias.alias_text)

    stmt = select(Einheit).where(Einheit.aktiv.is_(True)).order_by(Einheit.kuerzel.asc())
    return [{"kuerzel": einheit.kuerzel, "aliase": aliases_by_unit.get(einheit.id, [])} for einheit in db.scalars(stmt)]


def _build_prompt_group_context(db: Session) -> list[dict[str, object]]:
    parameter_names_by_id = {
        parameter.id: parameter.anzeigename
        for parameter in db.scalars(select(Laborparameter).where(Laborparameter.aktiv.is_(True)))
    }
    group_parameters: dict[str, list[str]] = {}
    assignment_stmt = select(GruppenParameter).order_by(GruppenParameter.sortierung.asc().nulls_last())
    for assignment in db.scalars(assignment_stmt):
        parameter_name = parameter_names_by_id.get(assignment.laborparameter_id)
        if parameter_name:
            group_parameters.setdefault(assignment.parameter_gruppe_id, []).append(parameter_name)

    stmt = select(ParameterGruppe).where(ParameterGruppe.aktiv.is_(True)).order_by(ParameterGruppe.name.asc())
    return [
        {
            "name": gruppe.name,
            "beschreibung": gruppe.beschreibung,
            "parameter": group_parameters.get(gruppe.id, []),
        }
        for gruppe in db.scalars(stmt)
    ]


def _delete_stored_document_file(dokument: Dokument) -> None:
    if not dokument.pfad_relativ:
        return

    root = get_documents_root()
    file_path = (root / dokument.pfad_relativ).resolve()
    try:
        file_path.relative_to(root)
    except ValueError as exc:
        raise ValueError("Der relative Dokumentpfad liegt außerhalb der konfigurierten Dokumentablage.") from exc

    if file_path.exists() and file_path.is_file():
        file_path.unlink()


def _format_prompt_type(prompt_typ: str) -> str:
    if prompt_typ == "tabelle":
        return "Tabelle/CSV/Excel"
    return "Laborbericht"


def _build_prompt_source_instruction(prompt_typ: str) -> str:
    if prompt_typ == "tabelle":
        return """Aufgabe:
Analysiere die bereitgestellte Tabelle, CSV- oder Excel-Datei vollständig und erzeuge daraus eine kurze Auswertung für den Anwender sowie ein gültiges JSON-Objekt nach dem Labordaten-Importvertrag V1.

Was mit der Quelle zu tun ist:
- Werte alle relevanten Tabellenblätter, Spalten, Überschriften, Einheiten- und Referenzspalten aus.
- Erkenne Spalten wie Parameter, Wert, Einheit, Referenzbereich, Entnahmedatum, Befunddatum, Labor, Kommentar und Gruppierung auch bei abweichender Benennung.
- Wenn mehrere Tabellenblätter oder Blöcke enthalten sind, überführe sie in eine gemeinsame Messwertliste und nutze erkennbare Blöcke als optionale "gruppenVorschlaege".
- Erfasse Laborwerte, Referenzwerte, Einheiten, qualitative Werte und fachlich relevante textliche Kommentare.
- Interpretiere die Werte nicht medizinisch. Es geht nur um strukturierte Datenerfassung."""

    return """Aufgabe:
Analysiere das angehängte Laborbericht-Dokument vollständig und erzeuge daraus eine kurze Auswertung für den Anwender sowie ein gültiges JSON-Objekt nach dem Labordaten-Importvertrag V1.

Was mit der Datei zu tun ist:
- Werte das komplette angehängte Dokument aus, nicht nur die erste sichtbare Tabelle.
- Berücksichtige Kopfbereich, Laborangaben, Person-/Auftragsangaben, Entnahmedatum, Befunddatum, Tabellen, Fußnoten, Methodenhinweise, Kommentare und erkennbare Berichtsblöcke.
- Erfasse Laborwerte, Referenzwerte, Einheiten, qualitative Werte und alle fachlich relevanten textlichen Ausführungen.
- Interpretiere den Befund nicht medizinisch. Es geht nur um strukturierte Datenerfassung."""


def _build_import_prompt_text(context_json: str, *, prompt_typ: str) -> str:
    source_instruction = _build_prompt_source_instruction(prompt_typ)
    return f"""Du bist ein Extraktionshelfer für die Anwendung "Labordaten".

{source_instruction}

Harte Ausgabevorgaben:
- Gib zuerst einen kurzen Überblick mit maximal 5 Stichpunkten aus.
- Nenne im Überblick die Zahl der erkannten Messwerte, fehlende oder unlesbare Pflichtangaben, unsichere Messwerte, nicht eindeutig gematchte Parameter und auffällige Widersprüche.
- Wenn keine Probleme erkennbar sind, schreibe das ausdrücklich kurz.
- Gib danach das Import-JSON in genau einem Markdown-Codeblock mit Sprache "json" aus, also ```json ... ```.
- Im JSON-Codeblock steht ausschließlich valides JSON ohne Kommentarzeilen.
- Verwende im JSON genau die Feldnamen aus dem Importvertrag.
- Erfinde keine Werte, Datumsangaben, Labore, Parameter, Referenzen, Einheiten oder IDs.
- Entferne Felder, für die es keinen Wert gibt. Setze keine Platzhalter wie "unbekannt".
- Zusätzliche Felder sind nicht erlaubt.
- Datumsformat ist immer YYYY-MM-DD.
- Zahlen im JSON verwenden Dezimalpunkt, auch wenn im Dokument ein Dezimalkomma steht.

Importvertrag V1:
- Wurzelobjekt: "schemaVersion", "quelleTyp", "personHinweis", "befund", "messwerte", optional "gruppenVorschlaege", optional "parameterVorschlaege".
- "schemaVersion" muss "1.0" sein.
- "quelleTyp" muss "ki_json" sein.
- "befund" braucht mindestens "entnahmedatum".
- "befund.personId" soll weggelassen werden, außer eine Person-ID wurde ausdrücklich in der Quelle oder durch den Nutzer vorgegeben. Die Person wird in der Anwendung beim Import ausgewählt oder überschrieben.
- "messwerte" ist eine Liste; jeder Eintrag braucht "originalParametername", "wertTyp" und "wertRohText".
- Erlaubte "wertTyp"-Werte: "numerisch", "text".
- Erlaubte "wertOperator"-Werte: "exakt", "kleiner_als", "kleiner_gleich", "groesser_als", "groesser_gleich", "ungefaehr".
- Erlaubte Referenz-Grenzoperatoren: "groesser_als", "groesser_gleich" für untere Grenzen; "kleiner_als", "kleiner_gleich" für obere Grenzen.
- Erlaubte Geschlechtscodes in Referenzen: "w", "m", "d"; sonst Feld weglassen.

Person:
- Setze "befund.personId" nicht, außer eine konkrete Person-ID wurde ausdrücklich mitgegeben.
- Wenn in der Quelle eine Person erkennbar ist, setze sie als Originaltext in "personHinweis".
- Die Anwendung kann beim Import eine Person auswählen oder überschreiben. Deshalb keine Person raten und keine Person aus vorhandenen Stammdaten ableiten.

Labor:
- Setze "befund.laborId" nur, wenn der Laborname im Dokument eindeutig zu einem bekannten Labor passt.
- Wenn kein eindeutiger Match möglich ist, setze "befund.laborName" mit dem erkannten Originalnamen.

Parameter und Werte:
- Setze "parameterId" nur bei eindeutigem Match auf einen bekannten Parameter, Anzeigenamen, internen Schlüssel oder Alias.
- Wenn der Match unsicher oder mehrdeutig ist, lasse "parameterId" weg, übernimm "originalParametername" exakt aus dem Dokument und setze "pruefbedarfFlag": true.
- "wertRohText" bleibt in der Originalschreibweise aus dem Dokument.
- Bei "<5" setze "wertOperator": "kleiner_als", "wertRohText": "<5" und "wertNum": 5.
- Bei ">200" setze "wertOperator": "groesser_als", "wertRohText": ">200" und "wertNum": 200.
- Bei qualitativen Werten wie "positiv", "negativ", "++", "nicht nachweisbar" nutze "wertTyp": "text" und fülle "wertText".
- Markiere unleserliche, widersprüchliche oder geschätzte Werte mit "unsicherFlag": true oder "pruefbedarfFlag": true und erkläre den Grund in "kiHinweis".
- Wenn ein Originalname offensichtlich nur eine alternative Schreibweise eines sicher gematchten Parameters ist, darf "aliasUebernehmen": true vorgeschlagen werden. Die Anwendung fragt später noch einmal nach.

Parameter-Vorschläge:
- Wenn ein Messwert nicht eindeutig zu einem bekannten Parameter passt, darfst Du ergänzend "parameterVorschlaege" anlegen.
- Ein Parameter-Vorschlag ist nur ein Vorschlag für die spätere Prüfung. Er ersetzt keinen sicheren Match und legt nichts automatisch an.
- Setze "anzeigename" als gut lesbaren Parameternamen, nicht nur als Abkürzung aus dem Dokument.
- Setze "messwertIndizes" mit allen Messwertpositionen, auf die sich der Vorschlag bezieht.
- Setze "wertTypStandard", "standardEinheit", "primaereKlassifikation", "beschreibungKurz", "moeglicheAliase" und "begruendungAusDokument" nur, wenn dies aus Dokument, üblicher Laborbezeichnung oder klarer Fachkenntnis belastbar ableitbar ist.
- Erlaubte "primaereKlassifikation"-Werte sind "krankwert", "schluesselwert" und "gesundmachwert". Diese Klassifikation beschreibt die typische Funktion des Parameters, nicht den konkreten Messwert.
- Orientierungsbeispiele: LDL-C, Small-LDLs, Triglyceride, HbA1c, HOMA-Index, Harnsäure, CRP/hsCRP, RANTES, oxidiertes LDL, TPO-AK, TAK, TRAK, ANA-AK, CCP-AK, Kreatinin, Cystatin C, GPT, GOT, gGT, CK, Beta-CrossLaps, TRAP 5b, ucOC, D-Ratio/Vitamin-D-Ratio, reverse T3, Bilirubin gesamt, Apo-B, Blei, Cadmium, DAO-Genetik, HNMT, I-FABP, IgE, Lipase, Lp(a), Lp-PLA2, MDA-LDL, Mikroalbuminurie, Nickel, Nitrotyrosin, NT-proBNP, Prolaktin im nicht-schwangeren Kontext und Zonulin sind typischerweise "krankwert"; eGFR, HDL-C, Homocystein, Ferritin, Transferrinsättigung, PTH, 1,25-OH-Vitamin-D, Calcium, AP/alkalische Phosphatase, BDNF, DHT, Estradiol/Östradiol, FSH, Gesamteiweiß, Histamin gesamt, Kalium, Kupfer, Leukozyten, LH, Natrium, Östron, Phosphat, Quick, Serotonin, SHBG, fT3, fT4, TSH, Thrombozyten und Ostase sind typischerweise "schluesselwert"; Magnesium, 25-OH-Vitamin-D, freies 25-OH-Vitamin-D, Bor, Mangan, B12/Holo-TC, Vitamin A, Vitamin B3/Nikotinamid, Vitamin C, Zink, Chrom, Selen, Jod, Eisen, DAO im Serum, SCFA, reduziertes Glutathion, Melatonin, Pregnenolon, Q10, Vitamin E, Alpha-Liponsäure, Lithium, Folsäure, Biotin, bioaktive B-Vitamine, Omega-3-Index, Molybdän, Aminosäureprofile, Progesteron und DHEA-S sind typischerweise "gesundmachwert". Bei Mehrfachangaben wie "S/G", "G/S", "K/S" oder "S/K" verwende eine plausible primäre Klassifikation und beschreibe die weitere Rolle in der Begründung.
- "beschreibungKurz" ist ausschließlich eine allgemeine, vom konkreten Bericht und Import unabhängige Fachbeschreibung: Was misst der Parameter oder wofür steht er typischerweise als Laborparameter?
- Schreibe in "beschreibungKurz" keine Sätze wie "Der Befund nennt...", "Im Bericht steht...", keine Methode aus diesem konkreten Dokument, keine Referenzbereiche, keine konkreten Werte, keine Diagnose und keine Bewertung des konkreten Messwerts.
- Wenn Du eine Anmerkung loswerden willst, warum der Vorschlag aus diesem Bericht abgeleitet wurde, schreibe sie ausschließlich in "begruendungAusDokument".
- "begruendungAusDokument" darf berichtsbezogen sein, z. B. Abschnitt, Methode, Einheit, Originalname oder warum mehrere Messwerte zu demselben Vorschlag gehören.
- Recherchiere nicht frei ins Blaue und erfinde keine Bedeutung. Wenn Du unsicher bist, setze "unsicherFlag": true oder lasse den Vorschlag weg.

Referenzen und Kommentare:
- Den originalen Referenztext immer in "referenzTextOriginal" erhalten, wenn vorhanden.
- Strukturierte "untereGrenzeNum" und "obereGrenzeNum" nur setzen, wenn die Grenzen eindeutig sind.
- Referenzeinheit in "referenzEinheit" setzen, wenn sie erkennbar ist.
- Alters- und geschlechtsbezogene Referenzkontexte in "referenzAlterMinTage", "referenzAlterMaxTage", "referenzGeschlechtCode" oder "referenzBemerkung" ablegen.
- Originale Labor-Kommentare zu einzelnen Werten in "bemerkungKurz" oder "bemerkungLang" übernehmen. Diese Felder sind für Text aus dem Bericht selbst gedacht, möglichst originalnah und ohne eigene KI-Erklärung.
- Eigene KI-Anmerkungen, Extraktionszweifel, Mapping-Hinweise oder Begründungen, warum ein Wert unsicher ist, gehören in "kiHinweis" und nicht in "bemerkungKurz" oder "bemerkungLang".

Gruppen:
- Wenn das Dokument erkennbare Berichtsabschnitte oder Blöcke enthält, erstelle "gruppenVorschlaege".
- "messwertIndizes" beziehen sich auf die Positionen im "messwerte"-Array, beginnend bei 0.
- Verwende Gruppen nur als Vorschläge; die Anwendung fragt später nach Übernahme, Zusammenführung oder Ignorieren.

Beispielstruktur:
{{
  "schemaVersion": "1.0",
  "quelleTyp": "ki_json",
  "personHinweis": "optional erkannte Person oder leer",
  "befund": {{
    "personId": "optional, nur wenn ausdrücklich vorgegeben",
    "laborId": "optional, nur bei sicherem Labor-Match",
    "laborName": "optional, wenn kein laborId-Match sicher ist",
    "entnahmedatum": "YYYY-MM-DD",
    "befunddatum": "YYYY-MM-DD",
    "bemerkung": "optional"
  }},
  "messwerte": [
    {{
      "parameterId": "optional, nur bei eindeutigem Parameter-Match",
      "originalParametername": "Name exakt aus dem Bericht",
      "wertTyp": "numerisch",
      "wertOperator": "exakt",
      "wertRohText": "41",
      "wertNum": 41,
      "einheitOriginal": "ng/ml",
      "referenzTextOriginal": "30-400 ng/ml",
      "untereGrenzeNum": 30,
      "obereGrenzeNum": 400,
      "referenzEinheit": "ng/ml",
      "bemerkungKurz": "optionaler Originalkommentar aus dem Laborbericht",
      "bemerkungLang": "optionaler längerer Originalkommentar aus dem Laborbericht",
      "kiHinweis": "optionale KI- oder Extraktionsanmerkung, nicht Teil des Laborbericht-Originaltexts",
      "aliasUebernehmen": false,
      "unsicherFlag": false,
      "pruefbedarfFlag": false
    }}
  ],
  "gruppenVorschlaege": [
    {{
      "name": "Berichtsabschnitt",
      "beschreibung": "optional",
      "messwertIndizes": [0]
    }}
  ],
  "parameterVorschlaege": [
    {{
      "anzeigename": "Gut lesbarer Parametername",
      "wertTypStandard": "numerisch",
      "standardEinheit": "optionale Einheit",
      "primaereKlassifikation": "krankwert | schluesselwert | gesundmachwert, optional",
      "beschreibungKurz": "Allgemeine, berichtsunabhängige Fachbeschreibung des Parameters",
      "moeglicheAliase": ["Name aus dem Bericht"],
      "begruendungAusDokument": "Berichtsbezogene Anmerkung, warum der Vorschlag zu den Messwerten passt",
      "unsicherFlag": false,
      "messwertIndizes": [0]
    }}
  ]
}}

Antwortformat:
1. Kurzer Überblick für den Anwender.
2. Danach ein einzelner ```json-Codeblock mit dem vollständigen Import-JSON. Der Codeblock ist die Kopiervorlage für die Anwendung.

Bekannte Stammdaten und Kontext:
{context_json}

Erzeuge jetzt aus der bereitgestellten Quelle den kurzen Überblick und danach das vollständige Import-JSON im json-Codeblock."""


def _build_detail(db: Session, importvorgang: Importvorgang) -> ImportvorgangDetailRead:
    payload = _parse_payload(importvorgang.roh_payload_text or "")
    checks = list_pruefpunkte(db, importvorgang.id)
    counts = _count_checks(db, importvorgang.id)
    parameter_resolver = ParameterResolver(db)
    imported_measurements = list(
        db.scalars(
            select(Messwert)
            .where(Messwert.importvorgang_id == importvorgang.id)
            .order_by(Messwert.erstellt_am.asc())
        )
    )
    imported_befund = db.scalar(
        select(Befund).where(Befund.importvorgang_id == importvorgang.id).order_by(Befund.erstellt_am.asc())
    )
    dokument_id = imported_befund.dokument_id if imported_befund is not None else importvorgang.dokument_id
    dokument = db.get(Dokument, dokument_id) if dokument_id else None
    gruppenvorschlaege = _build_group_suggestion_previews(
        db,
        payload=payload,
        imported_measurements=imported_measurements,
        parameter_resolver=parameter_resolver,
    )
    parameter_vorschlaege = _build_parameter_suggestion_previews(payload)
    parameter_suggestion_by_index = {
        messwert_index: suggestion
        for suggestion in parameter_vorschlaege
        for messwert_index in suggestion.messwert_indizes
    }

    return ImportvorgangDetailRead(
        id=importvorgang.id,
        quelle_typ=importvorgang.quelle_typ,
        status=importvorgang.status,
        person_id_vorschlag=importvorgang.person_id_vorschlag,
        schema_version=importvorgang.schema_version,
        bemerkung=importvorgang.bemerkung,
        warnungen_text=importvorgang.warnungen_text,
        fingerprint=importvorgang.fingerprint,
        dokument_id=dokument_id,
        dokument_dateiname=dokument.dateiname if dokument is not None else None,
        erstellt_am=importvorgang.erstellt_am,
        geaendert_am=importvorgang.geaendert_am,
        messwerte_anzahl=len(payload.messwerte),
        fehler_anzahl=counts["fehler"],
        warnung_anzahl=counts["warnung"],
        befund=ImportBefundPreviewRead(
            person_id=payload.befund.person_id,
            labor_id=imported_befund.labor_id if imported_befund is not None else payload.befund.labor_id,
            labor_name=payload.befund.labor_name,
            entnahmedatum=payload.befund.entnahmedatum,
            befunddatum=payload.befund.befunddatum,
            bemerkung=payload.befund.bemerkung,
            dokument_id=dokument_id,
            dokument_dateiname=dokument.dateiname if dokument is not None else None,
            dokument_pfad=(
                dokument.pfad_absolut
                if dokument is not None and dokument.pfad_absolut
                else payload.befund.dokument_pfad
            ),
        ),
        messwerte=[
            _build_measurement_preview(
                index=index,
                item=item,
                imported_measurements=imported_measurements,
                parameter_resolver=parameter_resolver,
                parameter_vorschlag=parameter_suggestion_by_index.get(index),
            )
            for index, item in enumerate(payload.messwerte)
        ],
        gruppenvorschlaege=gruppenvorschlaege,
        parameter_vorschlaege=parameter_vorschlaege,
        pruefpunkte=[ImportPruefpunktRead.model_validate(item) for item in checks],
    )


def _build_measurement_preview(
    *,
    index: int,
    item: ImportMesswertPayload,
    imported_measurements: list[Messwert],
    parameter_resolver: ParameterResolver,
    parameter_vorschlag: ImportParameterVorschlagRead | None = None,
) -> ImportMesswertPreviewRead:
    imported_parameter_id = imported_measurements[index].laborparameter_id if index < len(imported_measurements) else None
    resolution = parameter_resolver.resolve(
        original_name=item.original_parametername,
        explicit_parameter_id=item.parameter_id,
        manual_parameter_id=None,
    )

    if imported_parameter_id is not None:
        parameter_id = imported_parameter_id
        mapping_herkunft = "uebernommen"
        mapping_hinweis = parameter_resolver.get_parameter_name(imported_parameter_id)
    else:
        parameter_id = resolution.parameter_id
        mapping_herkunft = resolution.herkunft
        mapping_hinweis = resolution.hinweis

    return ImportMesswertPreviewRead(
        messwert_index=index,
        parameter_id=parameter_id,
        parameter_mapping_herkunft=mapping_herkunft,
        parameter_mapping_hinweis=mapping_hinweis,
        alias_uebernehmen=item.alias_uebernehmen,
        original_parametername=item.original_parametername,
        wert_typ=item.wert_typ,
        wert_operator=item.wert_operator,
        wert_roh_text=item.wert_roh_text,
        wert_num=item.wert_num,
        wert_text=item.wert_text,
        einheit_original=item.einheit_original,
        bemerkung_kurz=item.bemerkung_kurz,
        bemerkung_lang=item.bemerkung_lang,
        ki_hinweis=item.ki_hinweis,
        unsicher_flag=item.unsicher_flag,
        pruefbedarf_flag=item.pruefbedarf_flag,
        referenz_text_original=item.referenz_text_original,
        untere_grenze_num=item.untere_grenze_num,
        untere_grenze_operator=item.untere_grenze_operator,
        obere_grenze_num=item.obere_grenze_num,
        obere_grenze_operator=item.obere_grenze_operator,
        referenz_einheit=item.referenz_einheit,
        referenz_geschlecht_code=item.referenz_geschlecht_code,
        referenz_alter_min_tage=item.referenz_alter_min_tage,
        referenz_alter_max_tage=item.referenz_alter_max_tage,
        referenz_bemerkung=item.referenz_bemerkung,
        parameter_vorschlag=parameter_vorschlag,
    )


def _build_parameter_suggestion_previews(payload: ImportPayload) -> list[ImportParameterVorschlagRead]:
    return [
        ImportParameterVorschlagRead(
            index=index,
            anzeigename=suggestion.anzeigename,
            wert_typ_standard=suggestion.wert_typ_standard,
            standard_einheit=suggestion.standard_einheit,
            primaere_klassifikation=suggestion.primaere_klassifikation,
            beschreibung_kurz=suggestion.beschreibung_kurz,
            moegliche_aliase=suggestion.moegliche_aliase,
            begruendung_aus_dokument=suggestion.begruendung_aus_dokument,
            unsicher_flag=suggestion.unsicher_flag,
            messwert_indizes=suggestion.messwert_indizes,
        )
        for index, suggestion in enumerate(payload.parameter_vorschlaege)
    ]


def _build_parameter_suggestion_by_measurement_index(
    payload: ImportPayload,
) -> dict[int, ImportParameterVorschlagPayload]:
    suggestions_by_index: dict[int, ImportParameterVorschlagPayload] = {}
    for suggestion in payload.parameter_vorschlaege:
        for messwert_index in suggestion.messwert_indizes:
            suggestions_by_index.setdefault(messwert_index, suggestion)
    return suggestions_by_index


def _build_group_suggestion_previews(
    db: Session,
    *,
    payload: ImportPayload,
    imported_measurements: list[Messwert],
    parameter_resolver: ParameterResolver,
) -> list[ImportGruppenvorschlagRead]:
    existing_groups = _load_existing_group_contexts(db)
    previews: list[ImportGruppenvorschlagRead] = []

    for index, suggestion in enumerate(payload.gruppen_vorschlaege):
        parameter_ids, missing_indices = _resolve_group_suggestion_parameters(
            payload=payload,
            suggestion_index=index,
            imported_measurements=imported_measurements,
            parameter_resolver=parameter_resolver,
        )
        parameter_names = [
            parameter_resolver.get_parameter_name(parameter_id) or parameter_id
            for parameter_id in parameter_ids
        ]
        previews.append(
            ImportGruppenvorschlagRead(
                index=index,
                name=suggestion.name,
                beschreibung=suggestion.beschreibung,
                messwert_indizes=suggestion.messwert_indizes,
                parameter_ids=parameter_ids,
                parameter_namen=parameter_names,
                fehlende_messwert_indizes=missing_indices,
                aehnliche_gruppen=_find_similar_groups(
                    suggestion_name=suggestion.name,
                    resolved_parameter_ids=parameter_ids,
                    existing_groups=existing_groups,
                ),
                anwendbar=bool(parameter_ids) and not missing_indices,
            )
        )

    return previews


def _resolve_group_suggestion_parameters(
    *,
    payload: ImportPayload,
    suggestion_index: int,
    imported_measurements: list[Messwert],
    parameter_resolver: ParameterResolver,
) -> tuple[list[str], list[int]]:
    suggestion = payload.gruppen_vorschlaege[suggestion_index]
    resolved_parameter_ids: list[str] = []
    missing_indices: list[int] = []

    for messwert_index in suggestion.messwert_indizes:
        imported_parameter_id = (
            imported_measurements[messwert_index].laborparameter_id
            if 0 <= messwert_index < len(imported_measurements)
            else None
        )
        payload_item = payload.messwerte[messwert_index] if 0 <= messwert_index < len(payload.messwerte) else None
        if payload_item is None:
            missing_indices.append(messwert_index)
            continue

        if imported_parameter_id is not None:
            parameter_id = imported_parameter_id
        else:
            resolution = parameter_resolver.resolve(
                original_name=payload_item.original_parametername,
                explicit_parameter_id=payload_item.parameter_id,
                manual_parameter_id=None,
            )
            parameter_id = resolution.parameter_id

        if parameter_id is None:
            missing_indices.append(messwert_index)
            continue
        if parameter_id not in resolved_parameter_ids:
            resolved_parameter_ids.append(parameter_id)

    return resolved_parameter_ids, missing_indices


def _load_existing_group_contexts(db: Session) -> list[ExistingGroupContext]:
    stmt = (
        select(ParameterGruppe, GruppenParameter, Laborparameter)
        .outerjoin(GruppenParameter, GruppenParameter.parameter_gruppe_id == ParameterGruppe.id)
        .outerjoin(Laborparameter, GruppenParameter.laborparameter_id == Laborparameter.id)
        .where(ParameterGruppe.aktiv.is_(True))
        .order_by(ParameterGruppe.name.asc())
    )

    grouped: dict[str, ExistingGroupContext] = {}
    for gruppe, zuordnung, parameter in db.execute(stmt):
        context = grouped.setdefault(
            gruppe.id,
            ExistingGroupContext(
                gruppe_id=gruppe.id,
                name=gruppe.name,
                parameter_ids=set(),
                parameter_names_by_id={},
            ),
        )
        if zuordnung is not None and parameter is not None:
            context.parameter_ids.add(parameter.id)
            context.parameter_names_by_id[parameter.id] = parameter.anzeigename
    return list(grouped.values())


def _find_similar_groups(
    *,
    suggestion_name: str,
    resolved_parameter_ids: list[str],
    existing_groups: list[ExistingGroupContext],
) -> list[ImportAehnlicheGruppeRead]:
    normalized_name = normalize_parameter_name(suggestion_name)
    results: list[ImportAehnlicheGruppeRead] = []

    for group in existing_groups:
        overlap_ids = [parameter_id for parameter_id in resolved_parameter_ids if parameter_id in group.parameter_ids]
        name_match = normalized_name == normalize_parameter_name(group.name)
        if not overlap_ids and not name_match:
            continue

        results.append(
            ImportAehnlicheGruppeRead(
                gruppe_id=group.gruppe_id,
                name=group.name,
                parameter_anzahl=len(group.parameter_ids),
                gemeinsame_parameter_anzahl=len(overlap_ids),
                gemeinsame_parameter_namen=[group.parameter_names_by_id[parameter_id] for parameter_id in overlap_ids],
                namensaehnlich=name_match,
            )
        )

    results.sort(
        key=lambda item: (
            not item.namensaehnlich,
            -item.gemeinsame_parameter_anzahl,
            item.name.lower(),
        )
    )
    return results


def _find_active_group_by_name(db: Session, name: str) -> ParameterGruppe | None:
    normalized_name = name.strip().lower()
    if not normalized_name:
        return None
    return db.scalar(
        select(ParameterGruppe)
        .where(func.lower(ParameterGruppe.name) == normalized_name)
        .where(ParameterGruppe.aktiv.is_(True))
    )


def _generate_checks(
    db: Session,
    importvorgang_id: str,
    payload: ImportPayload,
    mappings: list[ImportParameterMapping],
) -> list[Pruefregel]:
    del importvorgang_id
    mapping_lookup = {item.messwert_index: item for item in mappings}
    checks: list[Pruefregel] = []
    parameter_resolver = ParameterResolver(db)

    person_id = payload.befund.person_id
    if not person_id:
        checks.append(
            Pruefregel("befund", "befund", "person_zuordnung", "fehler", "Dem Befund ist noch keine Person zugeordnet.")
        )
    elif db.get(Person, person_id) is None:
        checks.append(Pruefregel("befund", "befund", "person_zuordnung", "fehler", "Die zugeordnete Person existiert nicht."))
    else:
        person = db.get(Person, person_id)
        if (
            payload.person_hinweis
            and person is not None
            and not _person_hint_matches_context(payload.person_hinweis, person.anzeigename)
        ):
            checks.append(
                Pruefregel(
                    "befund",
                    "befund",
                    "person_hinweis",
                    "warnung",
                    f"Der im Dokument erkannte Personenhinweis '{payload.person_hinweis}' passt nicht eindeutig zur zugeordneten Person '{person.anzeigename}'.",
                )
            )

    if payload.befund.labor_id and db.get(Labor, payload.befund.labor_id) is None:
        checks.append(Pruefregel("befund", "befund", "labor_zuordnung", "fehler", "Das angegebene Labor existiert nicht."))

    duplicate_fingerprint = db.scalar(
        select(func.count(Importvorgang.id))
        .where(Importvorgang.fingerprint == hashlib.sha256(_serialize_payload(payload).encode("utf-8")).hexdigest())
        .where(Importvorgang.status != "verworfen")
    )
    if duplicate_fingerprint and duplicate_fingerprint > 1:
        checks.append(
            Pruefregel(
                "import",
                "gesamt",
                "fingerprint_duplikat",
                "warnung",
                "Ein sehr ähnlicher Importvorgang ist bereits in der Historie vorhanden.",
            )
        )

    for index, item in enumerate(payload.messwerte):
        mapping_request = mapping_lookup.get(index)
        resolution = (
            ParameterResolution(parameter_id=None, herkunft="neu")
            if _is_new_parameter_mapping(mapping_request)
            else parameter_resolver.resolve(
                original_name=item.original_parametername,
                explicit_parameter_id=item.parameter_id,
                manual_parameter_id=mapping_request.laborparameter_id if mapping_request is not None else None,
            )
        )
        parameter_id = resolution.parameter_id
        key = f"messwert:{index}"
        if _is_new_parameter_mapping(mapping_request):
            existing_resolution = parameter_resolver.resolve(
                original_name=item.original_parametername,
                explicit_parameter_id=item.parameter_id,
                manual_parameter_id=None,
            )
            if existing_resolution.parameter_id is not None:
                checks.append(
                    Pruefregel(
                        "messwert",
                        key,
                        "parameter_neuanlage",
                        "warnung",
                        f"Für '{item.original_parametername}' wurde Neuanlage gewählt, obwohl ein bestehender Parameter passt.",
                    )
                )
        elif parameter_id is None:
            meldung = (
                f"Für '{item.original_parametername}' gibt es mehrere passende Parameter"
                + (f": {resolution.hinweis}." if resolution.mehrdeutig and resolution.hinweis else ".")
                if resolution.mehrdeutig
                else f"Für '{item.original_parametername}' fehlt noch eine Parameterzuordnung."
            )
            checks.append(Pruefregel("messwert", key, "parameter_mapping", "warnung", meldung))
        else:
            parameter = db.get(Laborparameter, parameter_id)
            if parameter is None:
                checks.append(
                    Pruefregel(
                        "messwert",
                        key,
                        "parameter_mapping",
                        "fehler",
                        f"Der zugeordnete Parameter für '{item.original_parametername}' existiert nicht.",
                    )
                )
            elif parameter.wert_typ_standard != item.wert_typ:
                checks.append(
                    Pruefregel(
                        "messwert",
                        key,
                        "wert_typ_abgleich",
                        "warnung",
                        f"Der Werttyp von '{item.original_parametername}' passt nicht zum erwarteten Parametertyp.",
                    )
                )
            else:
                normalization_warning = _build_missing_normalization_warning(
                    db,
                    parameter=parameter,
                    item=item,
                )
                if normalization_warning is not None:
                    checks.append(
                        Pruefregel(
                            "messwert",
                            key,
                            "normierung_fehlt",
                            "warnung",
                            normalization_warning,
                        )
                    )

        alias_resolution = _resolve_alias_request(
            db,
            parameter_id=parameter_id,
            original_name=item.original_parametername,
            alias_requested=item.alias_uebernehmen or bool(mapping_request and mapping_request.alias_uebernehmen),
        )
        if alias_resolution.conflict_message:
            checks.append(
                Pruefregel(
                    "messwert",
                    key,
                    "alias_anlage",
                    "fehler",
                    alias_resolution.conflict_message,
                )
            )

        if item.wert_typ == "numerisch" and item.wert_num is None and not item.pruefbedarf_flag:
            checks.append(
                Pruefregel(
                    "messwert",
                    key,
                    "wert_parse",
                    "warnung",
                    f"Für '{item.original_parametername}' fehlt ein parsebarer Zahlenwert.",
                )
            )

        if person_id and parameter_id and _has_duplicate_measurement(db, person_id, payload.befund.entnahmedatum, parameter_id):
            checks.append(
                Pruefregel(
                    "messwert",
                    key,
                    "duplikat",
                    "warnung",
                    f"Für '{item.original_parametername}' gibt es am selben Entnahmedatum bereits einen Messwert.",
                )
            )

    return checks


def _person_hint_matches_context(person_hinweis: str, anzeigename: str) -> bool:
    normalized_hint = _normalize_person_hint(person_hinweis)
    normalized_name = _normalize_person_hint(anzeigename)
    if not normalized_hint or not normalized_name:
        return True
    if normalized_name in normalized_hint:
        return True

    name_tokens = {token for token in normalized_name.split() if len(token) >= 3}
    return any(token in normalized_hint for token in name_tokens)


def _normalize_person_hint(value: str) -> str:
    decomposed = unicodedata.normalize("NFKD", value)
    ascii_value = "".join(char for char in decomposed if not unicodedata.combining(char))
    return re.sub(r"[^a-z0-9]+", " ", ascii_value.casefold()).strip()


def _has_duplicate_measurement(db: Session, person_id: str, entnahmedatum, parameter_id: str) -> bool:
    stmt = (
        select(func.count(Messwert.id))
        .join(Befund, Messwert.befund_id == Befund.id)
        .where(Messwert.person_id == person_id)
        .where(Messwert.laborparameter_id == parameter_id)
        .where(Befund.entnahmedatum == entnahmedatum)
    )
    count = db.scalar(stmt) or 0
    return count > 0


def _build_missing_normalization_warning(
    db: Session,
    *,
    parameter: Laborparameter,
    item: ImportMesswertPayload,
) -> str | None:
    if item.wert_typ != "numerisch" or item.wert_num is None or not parameter.standard_einheit or not item.einheit_original:
        return None

    source_unit = _resolve_import_check_unit(db, item.einheit_original)
    target_unit = parameter.standard_einheit
    if source_unit == target_unit:
        return None

    normalized = parameter_conversions.resolve_measurement_normalization(
        db,
        laborparameter_id=parameter.id,
        wert_typ=item.wert_typ,
        wert_num=item.wert_num,
        einheit_original=source_unit,
    )
    if normalized.einheit_normiert == target_unit and normalized.wert_normiert_num is not None:
        return None

    return (
        f"Für '{item.original_parametername}' fehlt aktuell eine saubere Umrechnung von "
        f"'{source_unit}' in die führende Normeinheit '{target_unit}' des Parameters "
        f"'{parameter.anzeigename}'."
    )


def _resolve_import_check_unit(db: Session, unit: str) -> str:
    cleaned = einheiten_service.normalize_einheit(unit)
    if cleaned is None:
        return unit

    try:
        resolved = einheiten_service.require_existing_einheit(db, cleaned)
    except ValueError:
        return cleaned
    return resolved or cleaned


def _resolve_alias_request(
    db: Session,
    *,
    parameter_id: str | None,
    original_name: str,
    alias_requested: bool,
) -> AliasResolution:
    alias_text = original_name.strip()
    if not alias_requested or parameter_id is None or not alias_text:
        return AliasResolution(requested=False, alias_text=alias_text or None)

    parameter = db.get(Laborparameter, parameter_id)
    if parameter is None:
        return AliasResolution(
            requested=True,
            alias_text=alias_text,
            conflict_message=f"Der Zielparameter für den Alias '{alias_text}' existiert nicht.",
        )

    alias_normalized = normalize_parameter_name(alias_text)
    if not alias_normalized:
        return AliasResolution(requested=True, alias_text=alias_text)

    if alias_normalized in {
        normalize_parameter_name(parameter.anzeigename),
        normalize_parameter_name(parameter.interner_schluessel),
    }:
        return AliasResolution(requested=True, alias_text=alias_text, already_present=True)

    existing_alias = db.scalar(
        select(LaborparameterAlias).where(LaborparameterAlias.alias_normalisiert == alias_normalized)
    )
    if existing_alias is not None:
        if existing_alias.laborparameter_id == parameter.id:
            return AliasResolution(requested=True, alias_text=alias_text, already_present=True)
        other_parameter = db.get(Laborparameter, existing_alias.laborparameter_id)
        other_name = other_parameter.anzeigename if other_parameter is not None else "einem anderen Parameter"
        return AliasResolution(
            requested=True,
            alias_text=alias_text,
            conflict_message=f"Der Alias '{alias_text}' ist bereits dem Parameter '{other_name}' zugeordnet.",
        )

    for other_parameter in db.scalars(select(Laborparameter)):
        if other_parameter.id == parameter.id:
            continue
        if alias_normalized == normalize_parameter_name(other_parameter.anzeigename):
            return AliasResolution(
                requested=True,
                alias_text=alias_text,
                conflict_message=(
                    f"Der Alias '{alias_text}' kollidiert mit dem Anzeigenamen '{other_parameter.anzeigename}'."
                ),
            )
        if alias_normalized == normalize_parameter_name(other_parameter.interner_schluessel):
            return AliasResolution(
                requested=True,
                alias_text=alias_text,
                conflict_message=(
                    f"Der Alias '{alias_text}' kollidiert mit dem internen Schlüssel "
                    f"'{other_parameter.interner_schluessel}'."
                ),
            )

    return AliasResolution(requested=True, alias_text=alias_text)


def _is_new_parameter_mapping(mapping_request: ImportParameterMapping | None) -> bool:
    return bool(mapping_request and mapping_request.aktion == "neu")


def _create_parameter_from_import_mapping(
    db: Session,
    *,
    item: ImportMesswertPayload,
    mapping_request: ImportParameterMapping | None,
    standard_einheit: str | None,
    parameter_vorschlag: ImportParameterVorschlagPayload | None,
) -> str:
    anzeigename = (
        (mapping_request.neuer_parameter_name if mapping_request else None)
        or (parameter_vorschlag.anzeigename if parameter_vorschlag is not None else None)
        or item.original_parametername
    )
    anzeigename = anzeigename.strip()
    if not anzeigename:
        raise ValueError("Für einen neu anzulegenden Parameter fehlt der Name.")

    beschreibung = parameter_vorschlag.beschreibung_kurz if parameter_vorschlag is not None else None
    standard_einheit_value = (
        standard_einheit
        or (parameter_vorschlag.standard_einheit if parameter_vorschlag is not None else None)
    )
    wert_typ_standard = (
        parameter_vorschlag.wert_typ_standard
        if parameter_vorschlag is not None and parameter_vorschlag.wert_typ_standard
        else item.wert_typ
    )

    parameter = Laborparameter(
        interner_schluessel=_build_unique_import_parameter_key(db, anzeigename),
        anzeigename=anzeigename,
        beschreibung=beschreibung,
        standard_einheit=standard_einheit_value,
        wert_typ_standard=wert_typ_standard,
        primaere_klassifikation=parameter_vorschlag.primaere_klassifikation
        if parameter_vorschlag is not None
        else None,
    )
    db.add(parameter)
    db.flush()
    parameter_service.ensure_parameter_knowledge_page(db, parameter)
    return parameter.id


def _build_unique_import_parameter_key(db: Session, source_value: str | None) -> str:
    base_key = build_parameter_key_candidate(source_value)
    candidate = base_key
    suffix = 2
    while db.scalar(select(Laborparameter.id).where(Laborparameter.interner_schluessel == candidate)) is not None:
        candidate = f"{base_key}_{suffix}"
        suffix += 1
    return candidate


def _create_import_alias(
    db: Session,
    *,
    parameter_id: str,
    alias_text: str,
    bemerkung: str | None,
) -> None:
    alias_normalized = normalize_parameter_name(alias_text)
    if not alias_normalized:
        return

    db.add(
        LaborparameterAlias(
            laborparameter_id=parameter_id,
            alias_text=alias_text,
            alias_normalisiert=alias_normalized,
            bemerkung=bemerkung,
        )
    )
    db.flush()


def _store_checks(db: Session, importvorgang_id: str, checks: list[Pruefregel]) -> None:
    for item in checks:
        db.add(
            ImportPruefpunkt(
                importvorgang_id=importvorgang_id,
                objekt_typ=item.objekt_typ,
                objekt_schluessel_temp=item.objekt_schluessel_temp,
                pruefart=item.pruefart,
                status=item.status,
                meldung=item.meldung,
            )
        )


def _replace_checks(
    db: Session,
    importvorgang_id: str,
    checks: list[Pruefregel],
    bestaetige_warnungen: bool,
) -> None:
    for item in list_pruefpunkte(db, importvorgang_id):
        db.delete(item)
    for item in checks:
        is_warning = item.status == "warnung"
        db.add(
            ImportPruefpunkt(
                importvorgang_id=importvorgang_id,
                objekt_typ=item.objekt_typ,
                objekt_schluessel_temp=item.objekt_schluessel_temp,
                pruefart=item.pruefart,
                status="bestaetigt" if is_warning and bestaetige_warnungen else item.status,
                meldung=item.meldung,
                bestaetigt_vom_nutzer=is_warning and bestaetige_warnungen,
                bestaetigt_am=utcnow() if is_warning and bestaetige_warnungen else None,
            )
        )


def _count_checks(db: Session, import_id: str) -> dict[str, int]:
    stmt = (
        select(ImportPruefpunkt.status, func.count(ImportPruefpunkt.id))
        .where(ImportPruefpunkt.importvorgang_id == import_id)
        .group_by(ImportPruefpunkt.status)
    )
    result = {"fehler": 0, "warnung": 0}
    for status, count in db.execute(stmt):
        if status == "fehler":
            result["fehler"] = count
        elif status == "warnung":
            result["warnung"] = count
    return result


def _parse_payload(payload_json: str, person_id_override: str | None = None) -> ImportPayload:
    payload_json = _extract_import_json_text(payload_json)
    try:
        data = json.loads(payload_json)
    except json.JSONDecodeError as exc:
        raise ValueError(f"Das Import-JSON ist nicht gültig: {exc.msg}.") from exc

    if person_id_override:
        data.setdefault("befund", {})
        data["befund"]["personId"] = person_id_override

    try:
        return ImportPayload.model_validate(data)
    except Exception as exc:  # noqa: BLE001
        raise ValueError(f"Das Import-JSON passt nicht zum erwarteten Schema: {exc}") from exc


def _extract_import_json_text(raw_text: str) -> str:
    text = raw_text.strip()
    if text.startswith("{"):
        return text

    fenced_blocks = re.findall(r"```(?:json)?\s*(.*?)```", text, flags=re.IGNORECASE | re.DOTALL)
    for block in fenced_blocks:
        candidate = block.strip()
        if candidate.startswith("{"):
            return candidate

    first_brace = text.find("{")
    last_brace = text.rfind("}")
    if first_brace >= 0 and last_brace > first_brace:
        return text[first_brace : last_brace + 1]

    return text


def _serialize_payload(payload: ImportPayload) -> str:
    return json.dumps(payload.model_dump(mode="json", by_alias=True), ensure_ascii=False, sort_keys=True)


def _build_import_document_filename(
    db: Session,
    *,
    payload: ImportPayload,
    original_filename: str,
    override_name: str | None,
) -> str:
    original_path = Path(original_filename)
    extension = original_path.suffix or ".pdf"
    if override_name and override_name.strip():
        return _ensure_filename_extension(_sanitize_document_filename(override_name.strip()), extension)

    person_name = None
    if payload.befund.person_id:
        person = db.get(Person, payload.befund.person_id)
        person_name = person.anzeigename if person is not None else None

    labor_name = payload.befund.labor_name
    if not labor_name and payload.befund.labor_id:
        labor = db.get(Labor, payload.befund.labor_id)
        labor_name = labor.name if labor is not None else None

    date_value = payload.befund.entnahmedatum or payload.befund.befunddatum
    parts = [
        date_value.isoformat() if date_value else None,
        person_name,
        labor_name,
        "Laborbericht",
    ]
    filename = "_".join(_sanitize_document_filename(part) for part in parts if part)
    return _ensure_filename_extension(filename or original_path.stem or "Laborbericht", extension)


def _sanitize_document_filename(value: str) -> str:
    return re.sub(r"[^A-Za-z0-9ÄÖÜäöüß._-]+", "_", value).strip("._")


def _ensure_filename_extension(filename: str, extension: str) -> str:
    if Path(filename).suffix:
        return filename
    return f"{filename}{extension}"


def _resolve_labor(db: Session, payload: ImportPayload) -> str | None:
    if payload.befund.labor_id:
        return payload.befund.labor_id
    if payload.befund.labor_name:
        stmt = select(Labor).where(func.lower(Labor.name) == payload.befund.labor_name.lower())
        labor = db.scalar(stmt)
        if labor is not None:
            return labor.id
        labor = Labor(name=payload.befund.labor_name)
        db.add(labor)
        db.flush()
        return labor.id
    return None


def _summarize_warnings(checks: list[Pruefregel]) -> str | None:
    warnungen = [item.meldung for item in checks if item.status in {"warnung", "fehler"}]
    if not warnungen:
        return None
    return "\n".join(warnungen[:10])


def _parse_tabular_import(
    *,
    filename: str,
    content: bytes,
    person_id_override: str | None,
    labor_id_override: str | None,
    labor_name_override: str | None,
    entnahmedatum_override: date | None,
    befunddatum_override: date | None,
    befund_bemerkung_override: str | None,
    dokument: Dokument | None,
) -> ImportPayload:
    suffix = Path(filename).suffix.lower()
    rows = _read_tabular_rows(filename=filename, content=content, suffix=suffix)
    if not rows:
        raise ValueError("Die Datei enthält keine auswertbaren Zeilen.")

    entnahmedatum = entnahmedatum_override or _extract_date(rows, "entnahmedatum")
    if entnahmedatum is None:
        raise ValueError("Für CSV- oder Excel-Importe ist ein Entnahmedatum erforderlich.")

    befunddatum = befunddatum_override or _extract_date(rows, "befunddatum")
    befund_bemerkung = befund_bemerkung_override or _extract_text(rows, "befund_bemerkung")
    person_id = person_id_override or _extract_text(rows, "person_id")
    labor_id = labor_id_override or _extract_text(rows, "labor_id")
    labor_name = labor_name_override or _extract_text(rows, "labor_name")

    messwerte = [_row_to_measurement(row) for row in rows]
    messwerte = [item for item in messwerte if item is not None]
    if not messwerte:
        raise ValueError("Es konnten keine Messwerte aus der Datei abgeleitet werden.")
    gruppenvorschlaege = _build_tabular_group_suggestions(rows)

    quelle_typ = "excel" if suffix in {".xlsx", ".xlsm"} else "csv"
    dokument_pfad = dokument.pfad_absolut if dokument is not None else None

    return ImportPayload.model_validate(
        {
            "schemaVersion": "1.0",
            "quelleTyp": quelle_typ,
            "personHinweis": Path(filename).name,
            "befund": {
                "personId": person_id,
                "laborId": labor_id,
                "laborName": labor_name,
                "entnahmedatum": entnahmedatum.isoformat(),
                "befunddatum": befunddatum.isoformat() if befunddatum else None,
                "bemerkung": befund_bemerkung,
                "dokumentPfad": dokument_pfad,
            },
            "messwerte": [item.model_dump(mode="json", by_alias=True) for item in messwerte],
            "gruppenVorschlaege": gruppenvorschlaege,
        }
    )


def _read_tabular_rows(*, filename: str, content: bytes, suffix: str) -> list[dict[str, str]]:
    if suffix == ".csv":
        return _read_csv_rows(content)
    if suffix in {".xlsx", ".xlsm"}:
        return _read_excel_rows(content)
    raise ValueError(f"Der Dateityp von '{filename}' wird aktuell nur für CSV oder Excel unterstützt.")


def _read_csv_rows(content: bytes) -> list[dict[str, str]]:
    sample = content[:4096]
    try:
        text = content.decode("utf-8-sig")
    except UnicodeDecodeError:
        text = content.decode("latin-1")

    delimiter = ";"
    try:
        dialect = csv.Sniffer().sniff(sample.decode("utf-8-sig", errors="ignore"), delimiters=";,\t")
        delimiter = dialect.delimiter
    except csv.Error:
        delimiter = ";" if ";" in text else ","

    reader = csv.DictReader(io.StringIO(text), delimiter=delimiter)
    rows: list[dict[str, str]] = []
    for row in reader:
        normalized = {str(key or "").strip(): _stringify_cell(value) for key, value in row.items() if key}
        if any(value for value in normalized.values()):
            rows.append(normalized)
    return rows


def _read_excel_rows(content: bytes) -> list[dict[str, str]]:
    workbook = load_workbook(io.BytesIO(content), read_only=True, data_only=True)
    sheet = workbook.active
    iterator = sheet.iter_rows(values_only=True)
    try:
        header_row = next(iterator)
    except StopIteration as exc:
        raise ValueError("Die Excel-Datei enthält keine Daten.") from exc

    headers = [str(item).strip() if item is not None else "" for item in header_row]
    rows: list[dict[str, str]] = []
    for values in iterator:
        row = {headers[index]: _stringify_cell(value) for index, value in enumerate(values) if index < len(headers) and headers[index]}
        if any(value for value in row.values()):
            rows.append(row)
    workbook.close()
    return rows


def _row_to_measurement(row: dict[str, str]) -> ImportMesswertPayload | None:
    original_name = _row_value(row, "original_parametername")
    raw_value = _row_value(row, "wert_roh_text")
    explicit_num = _parse_optional_float(_row_value(row, "wert_num"))
    explicit_text = _row_value(row, "wert_text")

    if not original_name and not raw_value and explicit_num is None and not explicit_text:
        return None
    if not original_name:
        raise ValueError("Mindestens eine Zeile enthält keinen Parameternamen.")

    operator = _normalize_operator_code(_row_value(row, "wert_operator")) or "exakt"
    normalized_raw = raw_value or ""
    if raw_value:
        inferred_operator, cleaned_raw = _split_operator(raw_value)
        normalized_raw = cleaned_raw
        if operator == "exakt" and inferred_operator != "exakt":
            operator = inferred_operator

    raw_for_inference = normalized_raw or explicit_text or (str(explicit_num) if explicit_num is not None else "")
    wert_typ = _infer_value_type(_row_value(row, "wert_typ"), explicit_num, explicit_text, raw_for_inference)

    wert_num = explicit_num
    wert_text = explicit_text
    if wert_typ == "numerisch" and wert_num is None:
        wert_num = _parse_optional_float(normalized_raw)
    if wert_typ == "text" and not wert_text:
        wert_text = normalized_raw or raw_value or None

    reference_text = _row_value(row, "referenz_text_original")
    lower_reference_num = _parse_optional_float(_row_value(row, "untere_grenze_num"))
    upper_reference_num = _parse_optional_float(_row_value(row, "obere_grenze_num"))
    lower_reference_operator = _normalize_operator_code(_row_value(row, "untere_grenze_operator"))
    upper_reference_operator = _normalize_operator_code(_row_value(row, "obere_grenze_operator"))
    (
        lower_reference_operator,
        upper_reference_operator,
    ) = _infer_reference_operators(
        reference_text=reference_text,
        lower_value=lower_reference_num,
        upper_value=upper_reference_num,
        explicit_lower_operator=lower_reference_operator,
        explicit_upper_operator=upper_reference_operator,
    )

    payload = {
        "parameterId": _row_value(row, "parameter_id"),
        "originalParametername": original_name,
        "wertTyp": wert_typ,
        "wertOperator": operator,
        "wertRohText": raw_value or explicit_text or (str(explicit_num) if explicit_num is not None else ""),
        "wertNum": wert_num,
        "wertText": wert_text,
        "einheitOriginal": _row_value(row, "einheit_original"),
        "bemerkungKurz": _row_value(row, "bemerkung_kurz"),
        "bemerkungLang": _row_value(row, "bemerkung_lang"),
        "kiHinweis": _row_value(row, "ki_hinweis"),
        "referenzTextOriginal": reference_text,
        "untereGrenzeNum": lower_reference_num,
        "untereGrenzeOperator": lower_reference_operator,
        "obereGrenzeNum": upper_reference_num,
        "obereGrenzeOperator": upper_reference_operator,
        "referenzEinheit": _row_value(row, "referenz_einheit"),
        "referenzGeschlechtCode": _row_value(row, "referenz_geschlecht_code"),
        "referenzAlterMinTage": _parse_optional_int(_row_value(row, "referenz_alter_min_tage")),
        "referenzAlterMaxTage": _parse_optional_int(_row_value(row, "referenz_alter_max_tage")),
        "referenzBemerkung": _row_value(row, "referenz_bemerkung"),
        "aliasUebernehmen": _parse_bool(_row_value(row, "alias_uebernehmen")),
        "unsicherFlag": _parse_bool(_row_value(row, "unsicher_flag")),
        "pruefbedarfFlag": _parse_bool(_row_value(row, "pruefbedarf_flag")),
    }
    return ImportMesswertPayload.model_validate(payload)


def _extract_text(rows: list[dict[str, str]], field: str) -> str | None:
    for row in rows:
        value = _row_value(row, field)
        if value:
            return value
    return None


def _extract_date(rows: list[dict[str, str]], field: str) -> date | None:
    for row in rows:
        value = _parse_optional_date(_row_value(row, field))
        if value is not None:
            return value
    return None


def _build_tabular_group_suggestions(rows: list[dict[str, str]]) -> list[dict[str, object]]:
    grouped_indices: dict[str, list[int]] = {}
    group_metadata: dict[str, str | None] = {}

    for index, row in enumerate(rows):
        name = _row_value(row, "gruppe_name")
        if not name:
            continue
        cleaned_name = name.strip()
        if not cleaned_name:
            continue
        grouped_indices.setdefault(cleaned_name, []).append(index)
        group_metadata.setdefault(cleaned_name, _row_value(row, "gruppe_beschreibung"))

    return [
        {
            "name": name,
            "beschreibung": group_metadata[name],
            "messwertIndizes": indices,
        }
        for name, indices in grouped_indices.items()
    ]


def _row_value(row: dict[str, str], field: str) -> str | None:
    aliases = HEADER_ALIASES.get(field, {field})
    normalized_keys = {_normalize_header(key): value for key, value in row.items()}
    for alias in aliases:
        value = normalized_keys.get(alias)
        if value not in {None, ""}:
            return value
    return None


def _normalize_header(value: str) -> str:
    return re.sub(r"[^a-z0-9]+", "", value.lower())


def _stringify_cell(value) -> str:
    if value is None:
        return ""
    if isinstance(value, datetime):
        return value.date().isoformat()
    if isinstance(value, date):
        return value.isoformat()
    return str(value).strip()


def _parse_optional_float(value: str | None) -> float | None:
    if value is None:
        return None
    normalized = value.strip()
    if not normalized:
        return None
    normalized = normalized.replace(" ", "").replace("%", "")
    normalized = normalized.replace(".", "").replace(",", ".") if "," in normalized and "." in normalized and normalized.rfind(",") > normalized.rfind(".") else normalized.replace(",", ".")
    match = re.search(r"[-+]?\d+(?:\.\d+)?", normalized)
    if match is None:
        return None
    try:
        return float(match.group(0))
    except ValueError:
        return None


def _parse_optional_int(value: str | None) -> int | None:
    if value is None:
        return None
    normalized = value.strip()
    if not normalized:
        return None
    match = re.search(r"[-+]?\d+", normalized)
    if match is None:
        return None
    try:
        return int(match.group(0))
    except ValueError:
        return None


def _parse_optional_date(value: str | None) -> date | None:
    if value is None:
        return None
    normalized = value.strip()
    if not normalized:
        return None

    for fmt in ("%Y-%m-%d", "%d.%m.%Y", "%d/%m/%Y", "%d-%m-%Y"):
        try:
            return datetime.strptime(normalized, fmt).date()
        except ValueError:
            continue

    try:
        return datetime.fromisoformat(normalized).date()
    except ValueError:
        return None


def _parse_bool(value: str | None) -> bool:
    if value is None:
        return False
    normalized = value.strip().lower()
    return normalized in {"1", "true", "ja", "yes", "y", "x"}


def _split_operator(raw_value: str) -> tuple[str, str]:
    normalized = raw_value.strip()
    for token, code in (("<=", "kleiner_gleich"), (">=", "groesser_gleich"), ("<", "kleiner_als"), (">", "groesser_als")):
        if normalized.startswith(token):
            return code, normalized[len(token) :].strip()
    return "exakt", normalized


def _normalize_operator_code(value: str | None) -> str | None:
    if value is None:
        return None
    normalized = value.strip()
    return {
        "<": "kleiner_als",
        "<=": "kleiner_gleich",
        ">": "groesser_als",
        ">=": "groesser_gleich",
        "~": "ungefaehr",
    }.get(normalized, normalized or None)


def _infer_reference_operators(
    *,
    reference_text: str | None,
    lower_value: float | None,
    upper_value: float | None,
    explicit_lower_operator: str | None,
    explicit_upper_operator: str | None,
) -> tuple[str | None, str | None]:
    lower_operator = explicit_lower_operator or None
    upper_operator = explicit_upper_operator or None

    if reference_text:
        inferred_operator, _ = _split_operator(reference_text)
        if inferred_operator in {"kleiner_als", "kleiner_gleich"} and upper_value is not None and upper_operator is None:
            upper_operator = inferred_operator
        if inferred_operator in {"groesser_als", "groesser_gleich"} and lower_value is not None and lower_operator is None:
            lower_operator = inferred_operator

    if lower_value is not None and lower_operator is None:
        lower_operator = DEFAULT_LOWER_REFERENCE_OPERATOR
    if upper_value is not None and upper_operator is None:
        upper_operator = DEFAULT_UPPER_REFERENCE_OPERATOR

    return lower_operator, upper_operator


def _infer_value_type(
    explicit_type: str | None,
    explicit_num: float | None,
    explicit_text: str | None,
    raw_value: str | None,
) -> str:
    normalized_type = (explicit_type or "").strip().lower()
    if normalized_type in {"numerisch", "numeric", "zahl"}:
        return "numerisch"
    if normalized_type in {"text", "qualitativ", "qualitativtext"}:
        return "text"
    if explicit_num is not None:
        return "numerisch"
    if explicit_text:
        return "text"
    if raw_value and _parse_optional_float(raw_value) is not None:
        return "numerisch"
    return "text"
